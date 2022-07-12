import openpyxl
from random import randint as ri
from random import choice as ch
from numpy import *

loaded_data=loadtxt('data.csv',delimiter=',',dtype=float32)

x_data=loaded_data[:,0:-1]
t_data=loaded_data[:,[-1]]

W=random.rand(3,1)
b=random.rand(1)
print('W= ',W,', W.shape= ', W.shape,'b= ',b,', b.shape= ',b.shape)

def loss_func(x,t):
    y=dot(x,W)+b
    return (sum((t-y)**2))/(len(x))

def numerical_derivative(f,x):
    delta_x=1e-4
    grad=zeros_like(x)

    it=nditer(x,flags=['multi_index'],op_flags=['readwrite'])

    while not it.finished:
        idx=it.multi_index
        tmp_val=x[idx]
        x[idx]=float(tmp_val)+delta_x
        fx1=f(x)

        x[idx]=tmp_val-delta_x
        fx2=f(x)
        grad[idx]=(fx1-fx2)/(2*delta_x)

        x[idx]=tmp_val
        it.iternext()

    return grad

def error_val(x,t):
    y=dot(x,W)+b

    return (sum((t-y)**2))/(len(x))

def predict(x):
    y=dot(x,W)+b

    return y

learning_rate=1e-5

f=lambda x: loss_func(x_data,t_data)

print('Initial error value= ',error_val(x_data,t_data),', Initial W= ',W,', b= ',b)

for step in range(20001):
    W-=learning_rate*numerical_derivative(f,W)
    b-=learning_rate*numerical_derivative(f,b)

    if step%400==0:
        print('step= ',step,'error value= ',error_val(x_data,t_data),'W= ',W,', b= ',b)

wb=openpyxl.load_workbook('data.xlsx')
ws=wb.active

for i in range(8000):
    rn=ri(1,30)
    countob=[rn+ri(1,5),rn+ri(1,5),rn+ri(1,5)]
    c=[1]
    if True:
        if True:
            for i in c:
                if not ws.cell(row=i,column=1).value:
                    c=c[-1]
                    break
                c.append(c[-1]+1)
            for i in range(3):
                ws.cell(row=c,column=1+i).value=countob[i]
            ws.cell(row=c,column=4).value=round(predict(countob)[0])
wb.save('data.xlsx')
